import pandas as pd
import openpyxl
from module.InputExcelFile import InputExcelFile
from module.BulkInfo import BulkInfo
import numpy

class OutputExcelFile:
    """
    既存の一括登録シートをコピーして反映する
    output_file_name (str): 出力Excelファイルの名前
    output_sheet_name (str): 出力Excelファイルのシート名
    """
    def __init__(self, df,management_id,management_number,category,bulk_excel_file_path):
        self.data_list = self.df_to_list(df)
        self.bulk_excel_file_path = bulk_excel_file_path
        self.wb = self.load_bulk_excel()
        self.all_ws = self.wb['検査結果(ページ単位)']
        self.individual_ws = self.wb['検査結果(検査箇所単位)']
        self.copy_all_ws = self.copy_ws(self.all_ws)
        self.copy_individual_ws = self.copy_ws(self.individual_ws)
        self.merge_list = []

        self.main()
    #一括登録ファイルを取得
    def load_bulk_excel(self):
        wb = openpyxl.load_workbook(self.bulk_excel_file_path)
        return wb
    # 一括登録ファイルのシートをコピーする
    def copy_ws(self,ws):
        ws_copy = self.wb.copy_worksheet(ws)
        ws_copy.title = '_' + ws.title
        # コピーしたシートを表紙シートのあとに移動
        self.wb._sheets.remove(ws_copy)
        self.wb._sheets.append(ws_copy)
        return ws_copy

    # new_bulk_excel_file_pathにdfを書き込む
    def write_excel(self):
        #2行目から書き込む
        all_index = 2
        individual_index = 2
        for row in self.data_list:
            if(row[3] == 1):
                self.write_cell(self.all_ws,all_index,row)
                all_index += 1
            else:
                self.write_cell(self.individual_ws,individual_index,row)
                individual_index += 1
        self.wb.save(self.bulk_excel_file_path)
    # write_excelのサブモジュール
    # 引数にワークシートオブジェクト
    # ワークシートオブジェクトのセルに値を代入
    def write_cell(self,ws,index,data_row):

        result = data_row[2]
        _comment = data_row[4]
        _target = data_row[5]
        _ammend = data_row[6] 

        bulk_result = ws[index][3]
        bulk_comment = ws[index][5]
        bulk_target = ws[index][6]
        bulk_ammend = ws[index][7]
        #コメント列が結合セルか判定
        is_merged = self.is_merged_cell_not_top_left(bulk_target,ws)
        #検査結果は結合されないため分岐から除外
        bulk_result.value = result
        
        # いいえ以外コメント、対象ソースコード、修正ソースコードには何もしない
        if result != "いいえ":
            return

        #/rを削除
        comment = self.trim_r(_comment) if _comment is not None else ''
        target = self.trim_r(_target) if _target is not None else ''
        ammend = self.trim_r(_ammend) if _ammend is not None else ''

        if (not(is_merged)):
            # 結合セルではない場合はそのまま代入
            bulk_comment.value = comment
            bulk_target.value = target
            bulk_ammend.value = ammend
        elif(is_merged not in self.merge_list):
            # 反映をして、結合セルのリストに追加
            bulk_comment.value = comment
            bulk_target.value = target
            bulk_ammend.value = ammend
            self.merge_list.append(is_merged)
        #検査結果がいいえの結合セルの扱い
        else:
            #　結合セルであり、いいえの判定が複数ある場合
            # 検査項目を格納
            inspection = ws[is_merged][1].value
            row_number = ws[is_merged][4].value
            self.isUpdate(ws[is_merged][5],comment,inspection,row_number)
            self.isUpdate(ws[is_merged][6],target,inspection,row_number)
            self.isUpdate(ws[is_merged][7],ammend,inspection,row_number)

    def isUpdate(self,bulk_ws,value,inspection,row_number):
        if pd.isna(value):
            return
        if bulk_ws.value != value:
            print(f"""
検査項目：{inspection}
行番号:{row_number}
どちらを反映するか選択してください
1：{bulk_ws.value}
2：{value}
""")
            ans = input()
            if ans == '2':
                bulk_ws.value = value
    def trim_r(self,cell_value):
        # 値に改行コード(_x000D_)が含まれている場合は削除
        _value = cell_value
        if isinstance(cell_value, str):
            _value = cell_value.replace('_x000D_', '')
        return _value

    # 結合セルであり結合の最初のセルではない場合は最初の行番号を返す
    def is_merged_cell_not_top_left(self,cell, ws):
        for merged_range in ws.merged_cells.ranges:
            # cell.coordinateはセルのアドレス
            if cell.coordinate in merged_range:
                return merged_range.min_row  # 結合セルの場合は最初の行番号を返す
        return False  # 結合セルでない場合
    
    # dfをopenpyxlで扱えるようにリストに変換
    def df_to_list(self,df):
        data_list = df.values.tolist()
        return data_list

    def main(self):
        
        # Excelファイルの書き込み
        self.write_excel()

if __name__ == '__main__':
    bulk_path="C:/Users/user/OneDrive/プロジェクト/Python/Excel/test_data/本物1867_NUL0000_リンク_検査結果一括更新_20250101163437.xlsx"
    bulk_info = BulkInfo(bulk_path)

    # テスト用
    input_file_path = input('検査.xlsxファイルパスを入力してください')
    input_sheet_name = '検査'
    condition = {'管理番号':'NUL0000','検査カテゴリ':'リンク'}
    start_row = 3
    input_excel = InputExcelFile(input_file_path, input_sheet_name,condition,input_start_row=start_row)
    input_excel_file = InputExcelFile(input_file_path, input_sheet_name,condition,input_start_row=start_row)

    OutputExcelFile(input_excel_file.df,
                         bulk_info.management_id,
                         bulk_info.management_number,
                         bulk_info.category,
                         bulk_path
                         )
    