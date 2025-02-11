import openpyxl

class MergeCell:
    """
    一括登録シートの結合セルの扱いを行うクラス
    """
    def __init__(self, bulk_excel_file_path):
        self.bulk_excel_file_path = bulk_excel_file_path
        self.wb = self.load_bulk_excel()
        self.ws = self.wb['検査結果(検査箇所単位)']
        self.merge_cell_list = self.get_merge_cell_list()

    #一括登録ファイルを取得
    def load_bulk_excel(self):
        wb = openpyxl.load_workbook(self.bulk_excel_file_path)
        return wb
    
    # 結合セルのリストを取得
    def get_merge_cell_list(self):
        merge_cell_list = []
        for merge_cell in self.ws.merged_cells.ranges:
            #結合セルでも行数が一個、または行番号列は除外
            count_row = merge_cell.max_row - merge_cell.min_row + 1
            if count_row < 2 or merge_cell.min_col == 5:
                continue
            merge_cell_list.append(merge_cell)
        return merge_cell_list
    
    # 結合セルではない、左上セル、左上セルではないの３つを判定する
    def is_merged_cell(self, row,col):
        for merge_cell in self.merge_cell_list:
            if merge_cell.min_row <= row <= merge_cell.max_row and merge_cell.min_col <= col <= merge_cell.max_col:
                if row == merge_cell.min_row:
                    return 'top_left'
                else:
                    return 'not_top_left'
        return False
            
    
if __name__ == '__main__':
    bulk_path="C:/Users/user/OneDrive/プロジェクト/Python/Excel/test_data/本物1867_NUL0000_リンク_検査結果一括更新_20250101163437.xlsx"
    merge_cell = MergeCell(bulk_path)

    print(merge_cell.is_merged_cell(4,8))