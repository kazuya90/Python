import pandas as pd
import openpyxl

class BulkContent:
    def __init__(self, path,all_sheetname,individual_sheetname):
        self.path = path
        self.all_sheet = all_sheetname
        self.individual_sheet = individual_sheetname
        self.all_df = self.read_excel(self.all_sheet)
        self.individual_df = self.read_excel(self.individual_sheet)

        self.wb = openpyxl.load_workbook(self.path)
        self.all_ws = self.wb[self.all_sheet]
        self.individual_ws = self.wb[self.individual_sheet]

        self.all_merge_cells = self.detect_merge_cell(self.all_ws)
        self.individual_merge_cells = self.detect_merge_cell(self.individual_ws)
      
    def read_excel(self,_sheet_name):
        return pd.read_excel(self.path, sheet_name=_sheet_name,header=0)
    
    #結合したセルを検知する
    def detect_merge_cell(self,ws):
        merge_cells = ws.merged_cells.ranges
        #複数のセルが結合したもののみ（コロンが含まれるもの）を抽出
        merge_cells = [str(cell) for cell in merge_cells if ':' in str(cell)]
        return merge_cells

if __name__ == '__main__':
    path = 'C:/Users/user/OneDrive/プロジェクト/Python/Excel/test_data/1867_NUL0000_リンク_検査結果一括更新_20241227085441.xlsx'
    all_sheetname = '検査結果(ページ単位)'
    individual_sheetname = '検査結果(検査箇所単位)'
    bulk_content = BulkContent(path,all_sheetname,individual_sheetname)
    print(bulk_content.all_df)
    print(bulk_content.individual_df)
    print("全体",bulk_content.all_merge_cells)
    print("個別",bulk_content.individual_merge_cells)
    print('処理が完了しました。')